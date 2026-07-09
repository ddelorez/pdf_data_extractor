"""
Extraction Service - Business logic layer for PDF processing.
Wraps Phase 1 modules and manages job tracking and file lifecycle.
"""

import os
import uuid
import shutil
import json
import tempfile
import threading
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
from pathlib import Path
from typing import Dict, List, Any, Optional, Tuple
from datetime import datetime, timedelta
from enum import Enum

from openpyxl import Workbook
from openpyxl.styles import Font
from src.config import get_logger, START_ROW, COL_MAP, DEFAULT_TIMEOUT, DPR_DEFAULT_FORMAT
from src.core.pdf_processor import process_pdf
from src.core.excel_format_detector import (
    detect_excel_format,
    detect_dpr_format_key,
    ExcelFormat,
)
from src.core.excel_dpr_extraction import extract_dpr_records
from src.data.validator import validate_records
from src.data.deduplicator import deduplicate_and_sort
from src.data.dpr_qa import check_workbook_dates, check_month_gaps
from src.output.excel_writer import write_excel
from src.output.csv_writer import write_csv
from src.output.flowback_excel_writer import write_flowback_excel
from src.output.flowback_csv_writer import write_flowback_csv
from src.output.dpr_master_writer import write_dpr_master, merge_master

logger = get_logger(__name__)

# Maximum number of files allowed per batch (environment-configurable).
# This is a count cap only; the real byte ceiling is Flask's MAX_CONTENT_LENGTH
# (default 100 MB), which caps the whole multipart request — so all files in one
# request must collectively fit under it. See the size-limits note in app.py
# (issue #1.8).
MAX_BATCH_FILES = int(os.environ.get('MAX_BATCH_FILES', '50'))

# Bound the number of jobs processed concurrently in the background. Without a
# cap, each upload spawned its own thread (100 uploads -> 100 threads fighting
# over 2 CPU / 2 GB). Jobs beyond the cap queue and run as workers free up.
MAX_PROCESSING_WORKERS = int(os.environ.get('MAX_PROCESSING_WORKERS', '2'))

# Shared, module-level pool reused across all jobs (the service is per-process).
_processing_executor = ThreadPoolExecutor(
    max_workers=MAX_PROCESSING_WORKERS,
    thread_name_prefix="job-proc",
)

# How long (hours) a finished job's artifacts live before the sweeper deletes
# them, and how often the sweeper runs (seconds). Both env-configurable.
JOB_TTL_HOURS = float(os.environ.get('JOB_TTL_HOURS', '24'))
JOB_SWEEP_INTERVAL_SECONDS = float(os.environ.get('JOB_SWEEP_INTERVAL_SECONDS', '3600'))

# Bytes read from the head of an upload to confirm its real type.
_PDF_SNIFF_BYTES = 1024

# Accepted upload extensions mapped to a short "kind" tag used for on-disk naming
# and downstream routing (PDF pipeline vs DPR Excel->Excel pipeline).
_ACCEPTED_EXTENSIONS = {".pdf": "pdf", ".xlsx": "xlsx"}


def _looks_like_pdf(header: bytes) -> bool:
    """
    True if ``header`` (the first bytes of a file) is a PDF.

    Validates by magic bytes (%PDF-) rather than trusting the .pdf extension
    (issue #1.15). The marker is normally at offset 0, but the PDF spec and
    real-world readers tolerate a small amount of leading bytes, so we scan the
    first KB.
    """
    return b'%PDF-' in header[:_PDF_SNIFF_BYTES]


def _looks_like_xlsx(header: bytes) -> bool:
    """
    True if ``header`` looks like an .xlsx workbook.

    .xlsx is an Office Open XML file — a ZIP container — so it begins with the
    ZIP local-file-header magic ``PK\\x03\\x04`` (empty/spanned archives use
    ``PK\\x05\\x06`` / ``PK\\x07\\x08``). We check the ZIP signature by magic
    bytes rather than trusting the .xlsx extension, mirroring the PDF sniff.
    """
    return header[:2] == b'PK' and header[2:4] in (b'\x03\x04', b'\x05\x06', b'\x07\x08')


def _process_pdf_with_timeout(pdf_path: Path, timeout: float) -> List[Dict[str, Any]]:
    """
    Run process_pdf with a wall-clock timeout.

    A malformed PDF can hang pdfplumber forever and pin the processing thread.
    We run the extraction in a single-use worker and bail after ``timeout``
    seconds so the job is marked errored and the pool slot is freed.

    Soft-timeout limitation: on timeout the underlying worker thread is
    abandoned (Python cannot forcibly kill a thread), so a genuinely hung
    pdfplumber call keeps running in the background until it finishes. The
    bounded job pool (MAX_PROCESSING_WORKERS) caps how many such threads can
    accumulate. A hard kill would require a subprocess (see issue #1.2).
    """
    executor = ThreadPoolExecutor(max_workers=1, thread_name_prefix="pdf-extract")
    future = executor.submit(process_pdf, pdf_path)
    try:
        return future.result(timeout=timeout)
    except FuturesTimeoutError:
        raise ProcessingError(
            f"Processing timed out after {timeout:g}s: {pdf_path.name}"
        )
    finally:
        # Don't block on a possibly-hung worker; abandon it (documented above).
        executor.shutdown(wait=False)


class JobStatus(Enum):
    """Job processing status enumeration."""
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    ERROR = "error"
    CANCELLED = "cancelled"


class ProcessingError(Exception):
    """Custom exception for processing errors."""
    pass


class FileValidationError(Exception):
    """Custom exception for file validation errors."""
    pass


class NonPdfFileError(FileValidationError):
    """Raised when an upload is not a PDF (wrong extension or magic bytes).

    Subclasses FileValidationError so generic handlers still treat it as a
    validation failure, but the route layer maps it specifically to HTTP 418
    ("I'm a teapot") — this service brews PDFs, not whatever you sent it.
    """
    pass


class JobConflictError(ProcessingError):
    """Raised when a job is asked to process while not in a PENDING state.

    Subclasses ProcessingError so existing handlers still treat it as a known
    error, but the route layer maps it to HTTP 409 to prevent double-runs.
    """
    pass


class ProcessingJob:
    """
    Represents a processing job with state tracking.
    Manages file uploads, processing progress, and output files.
    """
    
    def __init__(self, job_id: str, upload_folder: str):
        """
        Initialize a processing job.
        
        Args:
            job_id: Unique job identifier (UUID)
            upload_folder: Base directory for file storage
        """
        self.job_id = job_id
        self.upload_folder = Path(upload_folder)
        self.job_folder = self.upload_folder / job_id
        self.job_folder.mkdir(parents=True, exist_ok=True)
        
        self.status = JobStatus.PENDING
        self.created_at = datetime.utcnow()
        self.completed_at: Optional[datetime] = None
        
        self.files_submitted: List[Path] = []
        self.records_extracted = 0
        self.records_valid = 0
        self.records_invalid = 0
        self.unique_wells = 0
        
        self.output_excel: Optional[Path] = None
        self.output_csv: Optional[Path] = None
        
        self.error_message: Optional[str] = None
        self.error_details: Optional[Dict[str, Any]] = None

        self.files_processed: int = 0
        self._cancel_requested: bool = False
        
        logger.info(f"Created ProcessingJob: {self.job_id}")
    
    def add_file(self, file_path: Path) -> None:
        """Add a file to the job."""
        if file_path.exists():
            self.files_submitted.append(file_path)
            logger.debug(f"Added file to job {self.job_id}: {file_path.name}")
    
    def set_processing(self) -> None:
        """Mark job as processing."""
        self.status = JobStatus.PROCESSING
        logger.info(f"Job {self.job_id} status: PROCESSING")
    
    def set_completed(self) -> None:
        """Mark job as completed."""
        self.status = JobStatus.COMPLETED
        self.completed_at = datetime.utcnow()
        logger.info(f"Job {self.job_id} status: COMPLETED")
    
    def set_error(self, message: str, details: Optional[Dict[str, Any]] = None) -> None:
        """Mark job as error."""
        self.status = JobStatus.ERROR
        self.error_message = message
        self.error_details = details or {}
        self.completed_at = datetime.utcnow()
        logger.error(f"Job {self.job_id} error: {message}")

    def request_cancel(self):
        """Request cancellation of this job."""
        self._cancel_requested = True
        self.status = JobStatus.CANCELLED

    @property
    def is_cancelled(self) -> bool:
        return self._cancel_requested

    def get_progress(self) -> int:
        """Get processing progress as percentage."""
        if self.status == JobStatus.COMPLETED:
            return 100
        elif self.status == JobStatus.ERROR:
            return 0
        elif self.status == JobStatus.CANCELLED:
            return self._calculate_file_progress()
        elif self.status == JobStatus.PROCESSING:
            return self._calculate_file_progress()
        else:
            return 0  # PENDING

    def _calculate_file_progress(self) -> int:
        """Calculate progress based on files processed."""
        total = len(self.files_submitted) if self.files_submitted else 0
        if total == 0:
            return 0
        # Reserve 10% for upload phase, 90% for processing
        return min(10 + int((self.files_processed / total) * 90), 99)
    
    def get_status_dict(self) -> Dict[str, Any]:
        """Get job status as dictionary."""
        return {
            "job_id": self.job_id,
            "status": self.status.value,
            "progress": self.get_progress(),
            "created_at": self.created_at.isoformat(),
            "completed_at": self.completed_at.isoformat() if self.completed_at else None,
            "files_submitted": len(self.files_submitted),
            "files_processed": self.files_processed,
            "records_extracted": self.records_extracted,
            "records_valid": self.records_valid,
            "records_invalid": self.records_invalid,
            "unique_wells": self.unique_wells,
            "records": self.records_valid,     # alias for frontend
            "wells": self.unique_wells,        # alias for frontend
        }
    
    def cleanup(self) -> None:
        """Clean up temporary files after download or expiration."""
        if self.job_folder.exists():
            try:
                shutil.rmtree(self.job_folder)
                logger.info(f"Cleaned up job folder: {self.job_folder}")
            except Exception as e:
                logger.warning(f"Failed to cleanup job folder {self.job_folder}: {e}")


class ExtractionService:
    """
    Service for managing PDF extraction workflows.
    Coordinates file uploads, processing, and output generation.
    """
    
    def __init__(self, upload_folder: str, template_folder: str, output_folder: str):
        """
        Initialize extraction service.
        
        Args:
            upload_folder: Directory for temporary file storage
            template_folder: Directory containing Excel templates
            output_folder: Directory for processed output files
        """
        self.upload_folder = Path(upload_folder)
        self.template_folder = Path(template_folder)
        self.output_folder = Path(output_folder)
        
        # Create directories if they don't exist
        self.upload_folder.mkdir(parents=True, exist_ok=True)
        self.template_folder.mkdir(parents=True, exist_ok=True)
        self.output_folder.mkdir(parents=True, exist_ok=True)

        # Job persistence directory (inside the volume-mounted output folder)
        self.jobs_dir = self.output_folder / "jobs"
        self.jobs_dir.mkdir(parents=True, exist_ok=True)

        # Thread safety for self.jobs dict access
        self._lock = threading.Lock()

        # Job tracking (in-memory cache)
        self.jobs: Dict[str, ProcessingJob] = {}

        # Per-file PDF processing timeout (seconds). Instance attribute so tests
        # can override it without touching the global default.
        self.pdf_timeout = DEFAULT_TIMEOUT

        # Background TTL sweeper state (started lazily via start_cleanup_sweeper)
        self._job_ttl_hours = JOB_TTL_HOURS
        self._sweep_interval = JOB_SWEEP_INTERVAL_SECONDS
        self._sweeper_thread: Optional[threading.Thread] = None
        self._sweeper_stop = threading.Event()

        # Restore jobs from disk on startup
        self._reload_jobs()

        logger.info("ExtractionService initialized")
        logger.info(f"  Upload folder: {self.upload_folder}")
        logger.info(f"  Template folder: {self.template_folder}")
        logger.info(f"  Output folder: {self.output_folder}")
        logger.info(f"  Jobs dir: {self.jobs_dir}")
    
    # ------------------------------------------------------------------
    # Job persistence helpers
    # ------------------------------------------------------------------

    def _persist_job(self, job: 'ProcessingJob') -> None:
        """
        Serialize a job's state to a JSON file atomically.

        Writes to a temporary file then renames to avoid partial writes
        from concurrent access or process crashes.
        """
        data = {
            "job_id": job.job_id,
            "upload_folder": str(job.upload_folder),
            "status": job.status.value,
            "created_at": job.created_at.isoformat(),
            "completed_at": job.completed_at.isoformat() if job.completed_at else None,
            "files_submitted": [str(p) for p in job.files_submitted],
            "records_extracted": job.records_extracted,
            "records_valid": job.records_valid,
            "records_invalid": job.records_invalid,
            "unique_wells": job.unique_wells,
            "output_excel": str(job.output_excel) if job.output_excel else None,
            "output_csv": str(job.output_csv) if job.output_csv else None,
            "error_message": job.error_message,
            "error_details": job.error_details,
            "files_processed": job.files_processed,
            "cancel_requested": job._cancel_requested,
        }
        job_file = self.jobs_dir / f"{job.job_id}.json"
        try:
            fd, tmp_path = tempfile.mkstemp(dir=str(self.jobs_dir), suffix=".tmp")
            try:
                with os.fdopen(fd, 'w') as f:
                    json.dump(data, f)
                os.replace(tmp_path, str(job_file))
            except Exception:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
                raise
        except Exception as e:
            logger.error(f"Failed to persist job {job.job_id}: {e}")

    def _load_job(self, job_id: str) -> Optional['ProcessingJob']:
        """
        Deserialize a job from its JSON file on disk.

        Returns a fully-restored ProcessingJob, or None if the file
        does not exist or cannot be parsed.
        """
        job_file = self.jobs_dir / f"{job_id}.json"
        if not job_file.exists():
            return None
        try:
            with open(str(job_file), 'r') as f:
                data = json.load(f)

            # Reconstruct object using the current upload folder so that
            # paths remain valid even if the container was recreated.
            job = ProcessingJob(job_id, str(self.upload_folder))
            job.status = JobStatus(data["status"])
            job.created_at = datetime.fromisoformat(data["created_at"])
            job.completed_at = (
                datetime.fromisoformat(data["completed_at"])
                if data.get("completed_at") else None
            )
            job.files_submitted = [Path(p) for p in data.get("files_submitted", [])]
            job.records_extracted = data.get("records_extracted", 0)
            job.records_valid = data.get("records_valid", 0)
            job.records_invalid = data.get("records_invalid", 0)
            job.unique_wells = data.get("unique_wells", 0)
            job.output_excel = Path(data["output_excel"]) if data.get("output_excel") else None
            job.output_csv = Path(data["output_csv"]) if data.get("output_csv") else None
            job.error_message = data.get("error_message")
            job.error_details = data.get("error_details")
            job.files_processed = data.get("files_processed", 0)
            job._cancel_requested = data.get("cancel_requested", False)
            return job
        except Exception as e:
            logger.error(f"Failed to load job {job_id} from disk: {e}")
            return None

    def _reload_jobs(self) -> None:
        """
        Scan the jobs directory on startup and restore all job files into
        the in-memory cache.  This allows the service to recover after a
        worker restart without losing previously submitted jobs.
        """
        count = 0
        try:
            for job_file in sorted(self.jobs_dir.glob("*.json")):
                job_id = job_file.stem
                job = self._load_job(job_id)
                if job is not None:
                    self.jobs[job_id] = job
                    count += 1
        except Exception as e:
            logger.error(f"Error during job reload: {e}")
        if count:
            logger.info(f"Restored {count} job(s) from disk on startup")

    def submit_files(self, files_list: List) -> str:
        """
        Submit files for extraction processing.
        
        Args:
            files_list: List of uploaded files (Werkzeug FileStorage objects)
        
        Returns:
            Job ID for tracking
        
        Raises:
            FileValidationError: If files are invalid
        """
        if not files_list:
            raise FileValidationError("No files provided")
        
        if len(files_list) > MAX_BATCH_FILES:
            raise FileValidationError(f"Maximum {MAX_BATCH_FILES} files per batch")
        
        # Create job
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id, str(self.upload_folder))
        
        # Save uploaded files
        saved_count = 0
        for file in files_list:
            if not file.filename:
                continue

            # Validate file extension against the accepted set (PDF for the
            # extraction pipeline, XLSX for the DPR Excel->Excel pipeline).
            ext = Path(file.filename).suffix.lower()
            kind = _ACCEPTED_EXTENSIONS.get(ext)
            if kind is None:
                raise NonPdfFileError(
                    f"Invalid file type: {file.filename}. "
                    "Only PDF and Excel (.xlsx) files are supported."
                )

            # Validate actual content by magic bytes, not just the extension
            # (issue #1.15). Peek at the header, then rewind so save() is intact.
            try:
                header = file.stream.read(_PDF_SNIFF_BYTES)
                file.stream.seek(0)
            except Exception as e:
                raise FileValidationError(f"Could not read file {file.filename}: {str(e)}")
            if kind == "pdf" and not _looks_like_pdf(header):
                raise NonPdfFileError(
                    f"Invalid PDF content: {file.filename} is not a valid PDF file."
                )
            if kind == "xlsx" and not _looks_like_xlsx(header):
                raise NonPdfFileError(
                    f"Invalid Excel content: {file.filename} is not a valid .xlsx file."
                )

            # Save file with UUID naming, preserving the real extension so the
            # background processor can route by modality.
            file_name = f"{uuid.uuid4().hex}.{kind}"
            file_path = job.job_folder / file_name

            try:
                file.save(str(file_path))
                job.add_file(file_path)
                saved_count += 1
                logger.info(f"Saved uploaded file: {file_name}")
            except Exception as e:
                raise FileValidationError(f"Failed to save file {file.filename}: {str(e)}")

        if saved_count == 0:
            raise FileValidationError("No valid PDF or Excel files were uploaded")

        # Store job and persist to disk (lock guards the in-memory dict)
        with self._lock:
            self.jobs[job_id] = job
        self._persist_job(job)

        # Auto-trigger background processing via the shared bounded pool.
        # Jobs beyond MAX_PROCESSING_WORKERS queue instead of spawning a thread
        # each, so a burst of uploads can't exhaust the CPU/memory budget.
        _processing_executor.submit(self._background_process, job_id)
        logger.info(f"Background processing queued for job {job_id}")

        logger.info(f"Job {job_id} submitted with {saved_count} files")
        return job_id

    def _background_process(self, job_id: str):
        """Run process_job in a background thread."""
        try:
            self.process_job(job_id)
        except Exception as e:
            logger.error(f"Background processing failed for job {job_id}: {e}")
            # process_job already handles setting error status
    
    def process_job(self, job_id: str, template_path: Optional[str] = None) -> Optional[Dict[str, Any]]:
        """
        Process uploaded PDFs for a job.
        
        Args:
            job_id: Job ID to process
            template_path: Optional path to Excel template
        
        Returns:
            Processing results dictionary
        
        Raises:
            ProcessingError: If processing fails
        """
        # Get job — fall back to disk if not in the in-memory cache.
        # The PENDING -> PROCESSING transition happens atomically under the lock
        # so the auto-trigger and a manual POST /api/process can't both run the
        # same job (which would corrupt counters and overwrite output files).
        with self._lock:
            if job_id not in self.jobs:
                loaded = self._load_job(job_id)
                if loaded is None:
                    raise ProcessingError(f"Job not found: {job_id}")
                self.jobs[job_id] = loaded
            job = self.jobs[job_id]

            if job.status != JobStatus.PENDING:
                raise JobConflictError(
                    f"Job {job_id} is not pending (status: {job.status.value}); "
                    "it has already been processed or is in progress"
                )

            if not job.files_submitted:
                raise ProcessingError("No files to process")

            job.set_processing()
        self._persist_job(job)

        try:

            # Route by input modality. The extension set was fixed at upload
            # time (submit_files only saves .pdf or .xlsx). A pure-xlsx batch is
            # the DPR Excel->Excel pipeline; a pure-pdf batch is the extraction
            # pipeline; anything else is a mixed/unsupported batch.
            exts = {p.suffix.lower() for p in job.files_submitted}
            if exts == {".xlsx"}:
                return self._process_dpr_job(job_id, job)
            if exts != {".pdf"}:
                raise ProcessingError(
                    "Mixed or unsupported file types in one job. Upload PDFs and "
                    "Excel (.xlsx) workbooks in separate batches."
                )

            # Process all PDFs and collect records
            all_records = []
            for pdf_path in job.files_submitted:
                # Check cancellation before processing each file
                if job.is_cancelled:
                    logger.info(f"Job {job_id} cancelled by user")
                    self._persist_job(job)
                    return
                try:
                    records = _process_pdf_with_timeout(pdf_path, self.pdf_timeout)
                    all_records.extend(records)
                    logger.info(f"Extracted {len(records)} records from {pdf_path.name}")
                except Exception as e:
                    logger.error(f"Failed to process {pdf_path.name}: {e}")
                    raise ProcessingError(f"Failed to process {pdf_path.name}: {str(e)}")
                job.records_extracted = len(all_records)
                job.files_processed += 1
                self._persist_job(job)  # Persist progress update
            
            job.records_extracted = len(all_records)
            
            if not all_records:
                raise ProcessingError("No records extracted from any PDF files")
            
            # Validate records
            valid_records, invalid_records = validate_records(all_records)
            job.records_valid = len(valid_records)
            job.records_invalid = len(invalid_records)
            
            if not valid_records:
                raise ProcessingError(
                    f"All {len(invalid_records)} extracted records failed validation"
                )
            
            logger.info(f"Validation: {len(valid_records)} valid, {len(invalid_records)} invalid")
            
            # Deduplicate and sort
            df = deduplicate_and_sort(valid_records)
            
            logger.info(f"After dedup: {len(df)} records")
            job.unique_wells = df['Well'].nunique() if 'Well' in df.columns else 0

            # Determine output format from _format tag on records
            format_tags = list(df['_format'].unique()) if '_format' in df.columns else ['narrative_sor']
            logger.info(f"Detected format tag(s): {format_tags}")

            if len(format_tags) > 1:
                raise ProcessingError(
                    "Mixed PDF formats detected. Please upload SOR and Flowback PDFs in separate batches."
                )

            detected_format = format_tags[0] if format_tags else 'narrative_sor'
            is_flowback = detected_format == 'tabular_flowback'

            excel_output_path = self.output_folder / f"{job_id}_output.xlsx"
            csv_output_path = self.output_folder / f"{job_id}_output.csv"

            if is_flowback:
                # Flowback format — use flowback-specific writers
                logger.info("Using flowback output writers (tabular_flowback format)")
                # Convert DataFrame back to list of dicts for flowback writers
                flowback_records: List[Dict[str, Any]] = [
                    {str(k): v for k, v in row.items()} for row in df.to_dict('records')
                ]
                try:
                    write_flowback_excel(flowback_records, excel_output_path)
                    job.output_excel = excel_output_path
                    logger.info(f"Flowback Excel file written: {excel_output_path}")
                except Exception as e:
                    raise ProcessingError(f"Failed to write flowback Excel file: {str(e)}")

                try:
                    write_flowback_csv(flowback_records, csv_output_path)
                    job.output_csv = csv_output_path
                    logger.info(f"Flowback CSV file written: {csv_output_path}")
                except Exception as e:
                    raise ProcessingError(f"Failed to write flowback CSV file: {str(e)}")
            else:
                # SOR / narrative format — existing logic
                logger.info("Using SOR output writers (narrative_sor format)")

                # Find template
                if template_path and Path(template_path).exists():
                    template = Path(template_path)
                else:
                    # Look for template.xlsx in project root
                    template = Path(__file__).parent.parent / "template.xlsx"

                # Write Excel output
                try:
                    if template.exists():
                        write_excel(df, template, excel_output_path)
                    else:
                        # Template not found — create Excel directly using openpyxl
                        logger.warning(
                            f"Template not found at {template}, creating Excel without template"
                        )
                        wb = Workbook()
                        ws = wb.active or wb.create_sheet("Sheet")
                        # Write bold headers in row 3 (one row above data START_ROW=4)
                        header_row = START_ROW - 1
                        _bold = Font(bold=True)
                        for _field_name, _col_num in COL_MAP.items():
                            _hcell = ws.cell(row=header_row, column=_col_num, value=_field_name)
                            _hcell.font = _bold
                        for i, row_data in enumerate(df.itertuples(index=False)):
                            excel_row = START_ROW + i
                            for field_name, col_number in COL_MAP.items():
                                if hasattr(row_data, field_name):
                                    value = getattr(row_data, field_name)
                                    if field_name == "Date" and value is not None:
                                        if not isinstance(value, datetime):
                                            try:
                                                value = datetime.combine(
                                                    value, datetime.min.time()
                                                )
                                            except Exception:
                                                value = None
                                    ws.cell(row=excel_row, column=col_number, value=value)
                        wb.save(str(excel_output_path))
                    job.output_excel = excel_output_path
                    logger.info(f"Excel file written: {excel_output_path}")
                except Exception as e:
                    raise ProcessingError(f"Failed to write Excel file: {str(e)}")

                # Write CSV output
                try:
                    write_csv(df, csv_output_path)
                    job.output_csv = csv_output_path
                    logger.info(f"CSV file written: {csv_output_path}")
                except Exception as e:
                    raise ProcessingError(f"Failed to write CSV file: {str(e)}")
            
            job.set_completed()
            self._persist_job(job)

            return {
                "status": "success",
                "job_id": job_id,
                "records": job.records_valid,
                "unique_wells": job.unique_wells,
                "invalid_records": job.records_invalid,
                "excel_url": f"/api/download/{job_id}/output.xlsx",
                "csv_url": f"/api/download/{job_id}/output.csv",
            }
        
        except ProcessingError as e:
            job.set_error(str(e))
            self._persist_job(job)
            raise
        except Exception as e:
            logger.exception(f"Unexpected error processing job {job_id}")
            job.set_error(
                "Unexpected error during processing",
                {"details": str(e), "type": type(e).__name__}
            )
            self._persist_job(job)
            raise ProcessingError(f"Unexpected error: {str(e)}")
    
    def _process_dpr_job(self, job_id: str, job: 'ProcessingJob') -> Dict[str, Any]:
        """Convert a batch of DPR Excel workbooks into an (appended) master.

        Classifies each uploaded .xlsx as either the existing master (at most
        one) or a raw monthly DPR (one or more), extracts records, appends to the
        master with (well, date) dedup (newest wins), and writes the combined
        master .xlsx plus a CSV mirror. QA flags (per-workbook date sanity +
        cross-month gaps) go to the master's "QA Flags" sheet.

        Runs inside process_job's try/except, so raising ProcessingError here is
        caught and recorded as job error by the caller.
        """
        # Classify uploaded workbooks by content.
        master_path: Optional[Path] = None
        raw_paths: List[Path] = []
        for path in job.files_submitted:
            fmt = detect_excel_format(path)
            if fmt == ExcelFormat.DPR_MASTER:
                if master_path is not None:
                    raise ProcessingError(
                        "Multiple master workbooks uploaded; include at most one "
                        "existing master to append to."
                    )
                master_path = path
            elif fmt == ExcelFormat.DPR_RAW:
                raw_paths.append(path)
            else:
                raise ProcessingError(
                    f"Unrecognized Excel workbook: {path.name}. Expected a "
                    "Walter Oil DPR workbook or an existing master."
                )

        if not raw_paths:
            raise ProcessingError(
                "No DPR workbook to convert; upload at least one monthly "
                "Walter Oil DPR .xlsx."
            )

        # Extract records and per-workbook date QA flags.
        all_records: List[Dict[str, Any]] = []
        qa_flags: List[Dict[str, str]] = []
        for path in raw_paths:
            if job.is_cancelled:
                logger.info(f"Job {job_id} cancelled by user")
                self._persist_job(job)
                return {"status": "cancelled", "job_id": job_id}
            fmt_key = detect_dpr_format_key(path) or DPR_DEFAULT_FORMAT
            try:
                recs = extract_dpr_records(path, format_key=fmt_key, source_name=path.name)
            except Exception as e:
                raise ProcessingError(f"Failed to extract {path.name}: {str(e)}")
            all_records.extend(recs)
            qa_flags.extend(
                check_workbook_dates(path, format_key=fmt_key, source_name=path.name)
            )
            job.files_processed += 1
            job.records_extracted = len(all_records)
            self._persist_job(job)

        # Count the master file toward files_processed for progress accuracy.
        if master_path is not None:
            job.files_processed += 1

        if not all_records:
            raise ProcessingError("No records extracted from the uploaded DPR workbook(s).")

        # Merge once to compute cross-month gap flags and summary counts, then
        # write the master with the complete QA set.
        merged = merge_master(all_records, existing_master_path=master_path)
        gap_flags = check_month_gaps(list(merged["date"])) if not merged.empty else []
        all_qa = qa_flags + gap_flags

        excel_output_path = self.output_folder / f"{job_id}_output.xlsx"
        csv_output_path = self.output_folder / f"{job_id}_output.csv"

        try:
            write_dpr_master(
                all_records,
                excel_output_path,
                existing_master_path=master_path,
                qa_flags=all_qa,
            )
            job.output_excel = excel_output_path
            logger.info(f"DPR master workbook written: {excel_output_path}")
        except Exception as e:
            raise ProcessingError(f"Failed to write master workbook: {str(e)}")

        try:
            merged.to_csv(csv_output_path, index=False)
            job.output_csv = csv_output_path
            logger.info(f"DPR CSV written: {csv_output_path}")
        except Exception as e:
            raise ProcessingError(f"Failed to write CSV file: {str(e)}")

        job.records_extracted = len(all_records)
        job.records_valid = len(merged)
        job.records_invalid = 0
        job.unique_wells = int(merged["well"].nunique()) if not merged.empty else 0

        job.set_completed()
        self._persist_job(job)
        return {
            "status": "success",
            "job_id": job_id,
            "records": job.records_valid,
            "unique_wells": job.unique_wells,
            "invalid_records": 0,
            "excel_url": f"/api/download/{job_id}/output.xlsx",
            "csv_url": f"/api/download/{job_id}/output.csv",
        }

    def get_job_status(self, job_id: str) -> Dict[str, Any]:
        """
        Get status of a processing job.
        
        Args:
            job_id: Job ID to check
        
        Returns:
            Status dictionary
        
        Raises:
            ProcessingError: If job not found
        """
        with self._lock:
            if job_id not in self.jobs:
                loaded = self._load_job(job_id)
                if loaded is None:
                    raise ProcessingError(f"Job not found: {job_id}")
                self.jobs[job_id] = loaded
            job = self.jobs[job_id]

        result = job.get_status_dict()
        
        if job.error_message:
            result["error"] = job.error_message
            result["error_details"] = job.error_details
        
        return result
    
    def cancel_job(self, job_id: str) -> dict:
        """Cancel a processing job."""
        with self._lock:
            job = self.jobs.get(job_id)
            if not job:
                job = self._load_job(job_id)
                if job:
                    self.jobs[job_id] = job

        if not job:
            raise ProcessingError(f"Job not found: {job_id}")

        if job.status in (JobStatus.COMPLETED, JobStatus.ERROR, JobStatus.CANCELLED):
            return job.get_status_dict()

        job.request_cancel()
        self._persist_job(job)
        logger.info(f"Job {job_id} cancellation requested")
        return job.get_status_dict()

    # ------------------------------------------------------------------
    # Artifact cleanup / garbage collection
    # ------------------------------------------------------------------

    def cleanup_job(self, job_id: str) -> None:
        """
        Delete all on-disk artifacts for a job and drop it from the cache.

        Removes the upload folder (uploads/<job_id>/), the output files
        (outputs/<job_id>_output.{xlsx,csv}), and the persisted job record
        (outputs/jobs/<job_id>.json). Idempotent — missing files are ignored.
        """
        with self._lock:
            job = self.jobs.pop(job_id, None)
        if job is None:
            job = self._load_job(job_id)

        # Upload folder (handled by ProcessingJob.cleanup)
        if job is not None:
            job.cleanup()

        # Output files
        for output_path in (
            self.output_folder / f"{job_id}_output.xlsx",
            self.output_folder / f"{job_id}_output.csv",
        ):
            try:
                output_path.unlink()
            except FileNotFoundError:
                pass
            except OSError as e:
                logger.warning(f"Failed to remove output {output_path}: {e}")

        # Persisted job record
        job_file = self.jobs_dir / f"{job_id}.json"
        try:
            job_file.unlink()
        except FileNotFoundError:
            pass
        except OSError as e:
            logger.warning(f"Failed to remove job record {job_file}: {e}")

        logger.info(f"Cleaned up artifacts for job {job_id}")

    def sweep_expired_jobs(self, now: Optional[datetime] = None) -> int:
        """
        Delete artifacts of finished jobs older than the configured TTL.

        Only terminal jobs (completed/error/cancelled) are eligible; in-flight
        and pending jobs are never swept. Age is measured from completed_at.
        Returns the number of jobs cleaned up.
        """
        now = now or datetime.utcnow()
        cutoff = now - timedelta(hours=self._job_ttl_hours)
        terminal = {JobStatus.COMPLETED, JobStatus.ERROR, JobStatus.CANCELLED}

        expired: List[str] = []
        for job_file in sorted(self.jobs_dir.glob("*.json")):
            job_id = job_file.stem
            job = self._load_job(job_id)
            if job is None or job.status not in terminal:
                continue
            finished_at = job.completed_at or job.created_at
            if finished_at < cutoff:
                expired.append(job_id)

        for job_id in expired:
            try:
                self.cleanup_job(job_id)
            except Exception as e:
                logger.error(f"Failed to sweep job {job_id}: {e}")

        if expired:
            logger.info(f"Sweeper cleaned up {len(expired)} expired job(s)")
        return len(expired)

    def start_cleanup_sweeper(self) -> None:
        """
        Start the background TTL sweeper thread (idempotent).

        Call once after the service is created (e.g. from the app factory in a
        non-testing context). With gunicorn --workers 1 there is a single
        service per process, so a single sweeper is sufficient.
        """
        if self._sweeper_thread is not None and self._sweeper_thread.is_alive():
            return
        self._sweeper_stop.clear()
        self._sweeper_thread = threading.Thread(
            target=self._sweeper_loop, name="job-sweeper", daemon=True
        )
        self._sweeper_thread.start()
        logger.info(
            f"Job cleanup sweeper started (ttl={self._job_ttl_hours}h, "
            f"interval={self._sweep_interval}s)"
        )

    def stop_cleanup_sweeper(self) -> None:
        """Signal the sweeper thread to stop (used in shutdown/tests)."""
        self._sweeper_stop.set()

    def _sweeper_loop(self) -> None:
        """Run sweep_expired_jobs every interval until stopped."""
        # Event.wait returns True when set (stop requested), False on timeout.
        while not self._sweeper_stop.wait(self._sweep_interval):
            try:
                self.sweep_expired_jobs()
            except Exception as e:
                logger.error(f"Cleanup sweeper error: {e}")

    def get_download_path(self, job_id: str, format: str) -> Path:
        """
        Get path to download file.
        
        Args:
            job_id: Job ID
            format: Output format ('xlsx' or 'csv')
        
        Returns:
            Path to output file
        
        Raises:
            ProcessingError: If file not found
        """
        with self._lock:
            if job_id not in self.jobs:
                loaded = self._load_job(job_id)
                if loaded is None:
                    raise ProcessingError(f"Job not found: {job_id}")
                self.jobs[job_id] = loaded
            job = self.jobs[job_id]

        if format.lower() == 'xlsx':
            if not job.output_excel or not job.output_excel.exists():
                raise ProcessingError("Excel output file not found")
            return job.output_excel
        elif format.lower() == 'csv':
            if not job.output_csv or not job.output_csv.exists():
                raise ProcessingError("CSV output file not found")
            return job.output_csv
        else:
            raise ProcessingError(f"Unsupported format: {format}")
