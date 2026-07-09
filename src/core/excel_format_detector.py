"""Excel workbook format detection module.

Classifies an uploaded ``.xlsx`` workbook as one of:

* a **partner DPR** raw input (e.g. Walter Oil monthly Daily Production Report)
  — recognised by the daily-sheet signature in :data:`DPR_EXCEL_FORMATS`, or
* an existing **master** workbook (the flat long-format output this pipeline
  appends to), recognised by its ``Data`` sheet header, or
* :data:`ExcelFormat.UNKNOWN`.

Mirrors the structure of :mod:`src.core.format_detector` (the PDF detector) but
reads cells with openpyxl instead of pdfplumber.
"""

from enum import Enum
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from src.config import (
    get_logger,
    DPR_EXCEL_FORMATS,
    DPR_MASTER_DATA_SHEET,
)

logger = get_logger(__name__)

# Value of the master ``Data`` sheet's first two header cells (A1, B1). Used to
# distinguish an already-converted master from a raw partner workbook.
_MASTER_HEADER_A1 = "well"
_MASTER_HEADER_B1 = "date"


class ExcelFormat(Enum):
    """Supported Excel workbook classifications."""

    DPR_RAW = "dpr_raw"        # a partner monthly DPR workbook (raw input)
    DPR_MASTER = "dpr_master"  # an existing combined master workbook
    UNKNOWN = "unknown"


def _cell(ws, coord: str):
    """Return the value at an A1-style ``coord``, or None on any failure."""
    try:
        return ws[coord].value
    except Exception:  # malformed coord / detached cell
        return None


def _sheet_matches_daily_signature(ws, fmt_cfg: dict) -> bool:
    """True if worksheet *ws* looks like a daily DPR sheet for *fmt_cfg*.

    Requires both the production-date label in ``date_label_cell`` and enough of
    the ``header_signature`` substrings in the configured header row. Requiring
    both keeps the ' Summary' and 'Meter totals' sheets (which have neither) out.
    """
    label = _cell(ws, fmt_cfg["date_label_cell"])
    if not isinstance(label, str) or fmt_cfg["date_label"].lower() not in label.lower():
        return False

    header_row = fmt_cfg["header_row"]
    try:
        row_values = [
            c.value for c in ws[header_row] if c.value is not None
        ]
    except Exception:
        return False
    combined = " ".join(str(v) for v in row_values).lower()

    hits = sum(1 for kw in fmt_cfg["header_signature"] if kw.lower() in combined)
    return hits >= len(fmt_cfg["header_signature"])


def _is_master(wb) -> bool:
    """True if workbook *wb* is an existing combined master (has a ``Data`` sheet
    whose header begins with the master ``well``/``date`` columns)."""
    if DPR_MASTER_DATA_SHEET not in wb.sheetnames:
        return False
    ws = wb[DPR_MASTER_DATA_SHEET]
    a1 = _cell(ws, "A1")
    b1 = _cell(ws, "B1")
    return (
        isinstance(a1, str) and a1.strip().lower() == _MASTER_HEADER_A1
        and isinstance(b1, str) and b1.strip().lower() == _MASTER_HEADER_B1
    )


def detect_excel_format(xlsx_path: Path) -> ExcelFormat:
    """Classify the workbook at *xlsx_path*.

    Returns :class:`ExcelFormat`. A raw partner DPR takes precedence over the
    master check (a raw workbook never contains a master ``Data`` sheet, but the
    ordering makes the intent explicit). Corrupt/unreadable workbooks resolve to
    :data:`ExcelFormat.UNKNOWN` rather than raising.
    """
    xlsx_path = Path(xlsx_path)
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except FileNotFoundError:
        raise
    except Exception as exc:
        logger.error("Could not open workbook %s: %s", xlsx_path.name, exc)
        return ExcelFormat.UNKNOWN

    try:
        for fmt_key, fmt_cfg in DPR_EXCEL_FORMATS.items():
            for ws in wb.worksheets:
                if _sheet_matches_daily_signature(ws, fmt_cfg):
                    logger.info(
                        "Detected DPR_RAW (%s) format: %s", fmt_key, xlsx_path.name
                    )
                    return ExcelFormat.DPR_RAW

        if _is_master(wb):
            logger.info("Detected DPR_MASTER format: %s", xlsx_path.name)
            return ExcelFormat.DPR_MASTER

        logger.info("Excel format UNKNOWN: %s", xlsx_path.name)
        return ExcelFormat.UNKNOWN
    finally:
        wb.close()


def detect_dpr_format_key(xlsx_path: Path) -> Optional[str]:
    """Return the :data:`DPR_EXCEL_FORMATS` key matching *xlsx_path*, or None.

    Convenience for the extractor, which needs the specific partner key (e.g.
    ``"walter_oil_dpr"``) rather than the coarse :class:`ExcelFormat`.
    """
    xlsx_path = Path(xlsx_path)
    try:
        wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    except FileNotFoundError:
        raise
    except Exception as exc:
        logger.error("Could not open workbook %s: %s", xlsx_path.name, exc)
        return None
    try:
        for fmt_key, fmt_cfg in DPR_EXCEL_FORMATS.items():
            for ws in wb.worksheets:
                if _sheet_matches_daily_signature(ws, fmt_cfg):
                    return fmt_key
        return None
    finally:
        wb.close()
