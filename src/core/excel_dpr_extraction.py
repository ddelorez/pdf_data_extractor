"""Extraction module for partner DPR Excel workbooks (Excel -> Excel pipeline).

Reshapes a monthly Daily Production Report workbook (e.g. Walter Oil's
``4-April DPR EW-834 2026.xlsx``) into flat per-well/per-day records suitable for
appending to the combined master. See plans/EXCEL_TO_EXCEL_DPR_SUPPORT.md.

Key rules encoded here:
  * Only integer-named ("daily") sheets are read; ' Summary' and 'Meter totals'
    are skipped.
  * The production date comes from ``date_cell`` (N4), NOT the sheet name — sheet
    "1" wraps to the last day of the month, so the sheet name is the *report* day.
  * Well rows live in fixed ranges with a blank gap row between well groups.
  * Non-numeric sentinels (S/I, N/A, NA) in numeric columns become blank (None);
    a genuine ``0`` in the source stays ``0``.
"""

from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from src.config import (
    get_logger,
    DPR_EXCEL_FORMATS,
    DPR_DEFAULT_FORMAT,
)

logger = get_logger(__name__)


def _is_daily_sheet_name(name: str) -> bool:
    """Daily sheets are named with a bare integer (possibly space-padded)."""
    return name.strip().isdigit()


def _clean_numeric(value: Any, sentinels: List[str]) -> Optional[float]:
    """Coerce a source cell to a number, mapping sentinels/blank to None.

    * ``None`` / empty string -> None
    * a configured sentinel (S/I, N/A, NA, case-insensitive) -> None
    * a number (int/float) -> float(value)  (full precision retained)
    * a numeric string -> float, else None (unrecognised text -> blank)
    """
    if value is None:
        return None
    if isinstance(value, bool):
        # Guard: bool is an int subclass; a stray TRUE/FALSE is not a measurement.
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if text == "":
        return None
    if text.upper() in {s.upper() for s in sentinels}:
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        logger.debug("Non-numeric value %r treated as blank", value)
        return None


def _normalize_date(value: Any) -> Optional[date]:
    """Return a ``date`` from a cell that may hold a datetime/date, else None."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _iter_well_rows(fmt_cfg: dict):
    """Yield 1-indexed worksheet row numbers holding well data."""
    for start, end in fmt_cfg["well_row_ranges"]:
        for row in range(start, end + 1):
            yield row


def extract_dpr_records(
    xlsx_path: Path,
    format_key: str = DPR_DEFAULT_FORMAT,
    source_name: Optional[str] = None,
) -> List[Dict[str, Any]]:
    """Extract flat per-well/per-day records from a DPR workbook.

    Args:
        xlsx_path: Path to the partner monthly DPR ``.xlsx``.
        format_key: Key into :data:`DPR_EXCEL_FORMATS` (partner layout).
        source_name: Label recorded on each record's ``_source`` (defaults to the
            workbook filename); used by QA reporting and provenance.

    Returns:
        A list of record dicts. Each has the real data fields
        (``well``, ``date``, ``Daily Oil``, ``Daily Gas``, ``Daily Water``,
        ``BHP``, ``FTP``, ``Choke Size``) plus provenance keys ``_sheet`` and
        ``_source``. Rows with no well id or no production date are skipped.

    Raises:
        ValueError: if *format_key* is unknown.
    """
    xlsx_path = Path(xlsx_path)
    if format_key not in DPR_EXCEL_FORMATS:
        raise ValueError(f"Unknown DPR format key: {format_key!r}")
    fmt_cfg = DPR_EXCEL_FORMATS[format_key]
    source = source_name or xlsx_path.name

    well_col = column_index_from_string(fmt_cfg["well_col"])
    # Pre-resolve output field -> column index once.
    col_idx = {
        field: column_index_from_string(letter)
        for field, letter in fmt_cfg["columns"].items()
    }
    date_cell = fmt_cfg["date_cell"]
    sentinels = fmt_cfg["sentinels"]

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    records: List[Dict[str, Any]] = []
    try:
        for sheet_name in wb.sheetnames:
            if not _is_daily_sheet_name(sheet_name):
                continue
            ws = wb[sheet_name]

            prod_date = _normalize_date(ws[date_cell].value)
            if prod_date is None:
                logger.warning(
                    "Sheet %r in %s has no valid production date at %s; skipping",
                    sheet_name, source, date_cell,
                )
                continue

            for row in _iter_well_rows(fmt_cfg):
                well = ws.cell(row=row, column=well_col).value
                if well is None or str(well).strip() == "":
                    continue  # blank gap row or missing well id

                record: Dict[str, Any] = {
                    "well": str(well).strip(),
                    "date": prod_date,
                    "_sheet": sheet_name.strip(),
                    "_source": source,
                }
                for field, cidx in col_idx.items():
                    record[field] = _clean_numeric(
                        ws.cell(row=row, column=cidx).value, sentinels
                    )
                records.append(record)
    finally:
        wb.close()

    logger.info(
        "Extracted %d record(s) from %s (%s)", len(records), source, format_key
    )
    return records
