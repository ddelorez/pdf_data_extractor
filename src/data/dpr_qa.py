"""Data-quality checks for the DPR Excel -> Excel pipeline.

Produces the rows that populate the master's ``QA Flags`` sheet
(columns ``Workbook | Sheet | Concern``). Two rule families, both
reverse-engineered from the partner's reference master:

1. **N4 date sanity** (:func:`check_workbook_dates`) — within one monthly
   workbook the daily sheets should all share a calendar month. The modal
   (most common) year-month is taken as expected; any sheet whose production
   date (cell N4) falls outside it, or is missing, is flagged. This reproduces
   sample flags like ``"N4 date was 2023-12-31"`` (stale template default) and
   ``"N4 date was 2026-03-01"`` (a February workbook whose wrap sheet held a
   March date).

2. **Cross-month gaps** (:func:`check_month_gaps`) — after appending, any
   calendar month between the earliest and latest production date with zero
   rows is flagged as a missing workbook.
"""

from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook

from src.config import get_logger, DPR_EXCEL_FORMATS, DPR_DEFAULT_FORMAT

logger = get_logger(__name__)


def _flag(workbook: str, sheet: str, concern: str) -> Dict[str, str]:
    """Build one QA-Flags row."""
    return {"Workbook": workbook, "Sheet": sheet, "Concern": concern}


def _as_date(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def check_workbook_dates(
    xlsx_path: Path,
    format_key: str = DPR_DEFAULT_FORMAT,
    source_name: Optional[str] = None,
) -> List[Dict[str, str]]:
    """Flag daily sheets whose N4 production date is missing or out-of-month.

    Args:
        xlsx_path: Path to the partner monthly DPR workbook.
        format_key: Key into :data:`DPR_EXCEL_FORMATS`.
        source_name: Label used in the ``Workbook`` column (defaults to filename).

    Returns:
        A list of QA-Flags rows (possibly empty).
    """
    xlsx_path = Path(xlsx_path)
    fmt_cfg = DPR_EXCEL_FORMATS[format_key]
    date_cell = fmt_cfg["date_cell"]
    source = source_name or xlsx_path.name

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        per_sheet: List[tuple] = []  # (sheet_name, date_or_None, raw_value)
        for sheet_name in wb.sheetnames:
            if not sheet_name.strip().isdigit():
                continue
            raw = wb[sheet_name][date_cell].value
            per_sheet.append((sheet_name.strip(), _as_date(raw), raw))
    finally:
        wb.close()

    if not per_sheet:
        return []

    # Expected month = modal (year, month) among the valid dates.
    months = Counter(
        (d.year, d.month) for _, d, _ in per_sheet if d is not None
    )
    flags: List[Dict[str, str]] = []
    if not months:
        # No sheet had a usable date at all.
        for name, _, raw in per_sheet:
            flags.append(_flag(source, name, f"{date_cell} has no valid date (was {raw!r})"))
        return flags

    expected = months.most_common(1)[0][0]
    for name, d, raw in per_sheet:
        if d is None:
            flags.append(_flag(source, name, f"{date_cell} has no valid date (was {raw!r})"))
        elif (d.year, d.month) != expected:
            flags.append(_flag(source, name, f"{date_cell} date was {d.isoformat()}"))

    if flags:
        logger.info("Workbook %s produced %d QA date flag(s)", source, len(flags))
    return flags


def check_month_gaps(
    dates: Iterable[date],
    source_label: str = "Uploaded workbooks",
) -> List[Dict[str, str]]:
    """Flag calendar months with no data between the min and max date.

    Args:
        dates: All production dates present in the combined master.
        source_label: Value placed in the ``Workbook`` column of gap flags.

    Returns:
        One QA-Flags row per missing month, chronologically ordered.
    """
    present = {(d.year, d.month) for d in dates if d is not None}
    if len(present) < 2:
        return []

    lo = min(present)
    hi = max(present)
    flags: List[Dict[str, str]] = []
    year, month = lo
    while (year, month) <= hi:
        if (year, month) not in present:
            ym = f"{year:04d}-{month:02d}"
            flags.append(_flag(source_label, ym, f"No uploaded DPR workbook for {ym}"))
        month += 1
        if month > 12:
            month = 1
            year += 1
    if flags:
        logger.info("Detected %d month-gap flag(s)", len(flags))
    return flags
