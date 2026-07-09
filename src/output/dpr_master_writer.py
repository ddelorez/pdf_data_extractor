"""Writer for the combined DPR master workbook (Excel -> Excel pipeline).

Merges freshly-extracted records into an (optional) existing master, applying the
locked policies:

  * **append** to the existing master rather than replacing it,
  * **dedup on (well, date)** keeping the *newest upload* (incoming wins),
  * **full decimal precision** (values are never rounded),
  * output columns follow :data:`DPR_MASTER_COLUMNS`, including the three blank
    spacer columns carried over from the partner's template.

Two sheets are written: ``Data`` (the long-format table) and ``QA Flags``.
"""

import os
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from src.config import (
    get_logger,
    DPR_MASTER_COLUMNS,
    DPR_RECORD_FIELDS,
    DPR_MASTER_DATA_SHEET,
    DPR_MASTER_QA_SHEET,
    DPR_QA_COLUMNS,
)

logger = get_logger(__name__)


def _coerce_date(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None or value == "":
        return None
    try:
        return pd.to_datetime(value).date()
    except Exception:
        return None


class MasterReadError(Exception):
    """Raised when an existing master workbook cannot be read.

    Deliberately fails loud instead of degrading to an empty master: silently
    returning no rows here would let a job report success while dropping the
    entire prior history (review finding HIGH-2).
    """
    pass


def load_existing_master(
    master_path: Path,
    qa_flags: Optional[List[Dict[str, str]]] = None,
) -> pd.DataFrame:
    """Read an existing master's ``Data`` sheet into the internal record schema.

    Selects only the real (non-spacer) columns by header name, tolerating the
    ``Unnamed: N`` spacer columns pandas assigns to the blank template columns.

    Fails loud (raises :class:`MasterReadError`) if the workbook cannot be read,
    rather than silently discarding history. Rows whose well/date cannot be
    parsed are excluded, and — because that also silently shrinks history — the
    count is logged and, when *qa_flags* is provided, surfaced as a QA flag.
    """
    master_path = Path(master_path)
    try:
        raw = pd.read_excel(master_path, sheet_name=DPR_MASTER_DATA_SHEET)
    except Exception as exc:
        raise MasterReadError(
            f"Could not read existing master {master_path.name}: {exc}"
        ) from exc

    # Keep only known fields; missing ones become all-NA columns.
    present = [c for c in DPR_RECORD_FIELDS if c in raw.columns]
    df = raw[present].copy()
    for field in DPR_RECORD_FIELDS:
        if field not in df.columns:
            df[field] = pd.NA
    df = df[DPR_RECORD_FIELDS]
    df["date"] = df["date"].map(_coerce_date)
    n_before = len(df)
    df = df[df["well"].notna() & df["date"].notna()]
    n_dropped = n_before - len(df)
    df["well"] = df["well"].astype(str).str.strip()

    if n_dropped:
        msg = (
            f"{n_dropped} existing master row(s) had an unparseable well/date "
            "and were excluded from the merged output"
        )
        logger.warning("%s (%s)", msg, master_path.name)
        if qa_flags is not None:
            qa_flags.append({
                "Workbook": master_path.name,
                "Sheet": DPR_MASTER_DATA_SHEET,
                "Concern": msg,
            })
    logger.info("Loaded %d existing master row(s) from %s", len(df), master_path.name)
    return df


def _records_to_df(records: List[Dict[str, Any]]) -> pd.DataFrame:
    """Project extracted records onto the internal record schema."""
    if not records:
        return pd.DataFrame(columns=DPR_RECORD_FIELDS)
    df = pd.DataFrame(records)
    for field in DPR_RECORD_FIELDS:
        if field not in df.columns:
            df[field] = pd.NA
    df = df[DPR_RECORD_FIELDS].copy()
    df["date"] = df["date"].map(_coerce_date)
    df = df[df["well"].notna() & df["date"].notna()]
    df["well"] = df["well"].astype(str).str.strip()
    return df


def merge_master(
    new_records: List[Dict[str, Any]],
    existing_master_path: Optional[Path] = None,
    qa_flags: Optional[List[Dict[str, str]]] = None,
) -> pd.DataFrame:
    """Combine new records with an optional existing master.

    Ordering guarantees the dedup keeps the newest upload: existing rows first,
    then incoming, then ``drop_duplicates(keep="last")`` on ``(well, date)``.
    Result is sorted by ``date`` then ``well``. Any QA flags raised while reading
    the existing master (dropped rows) are appended to *qa_flags* if provided.
    """
    frames: List[pd.DataFrame] = []
    if existing_master_path is not None:
        frames.append(load_existing_master(existing_master_path, qa_flags=qa_flags))
    frames.append(_records_to_df(new_records))

    combined = pd.concat(frames, ignore_index=True)
    if combined.empty:
        return combined

    before = len(combined)
    combined = combined.drop_duplicates(subset=["well", "date"], keep="last")
    dropped = before - len(combined)
    if dropped:
        logger.info("Dedup removed %d overlapping (well, date) row(s)", dropped)

    combined = combined.sort_values(by=["date", "well"], kind="stable").reset_index(drop=True)
    return combined


def write_master_dataframe(
    merged: pd.DataFrame,
    output_path: Path,
    qa_flags: Optional[List[Dict[str, str]]] = None,
) -> None:
    """Write an already-merged ``Data`` DataFrame + QA Flags to *output_path*.

    Separated from merging so callers that need the merged frame for other
    purposes (gap detection, CSV mirror, counts) can merge once and write once
    (review finding MED-4). The file is written atomically via a temp file +
    ``os.replace`` so a crash mid-save never leaves a truncated workbook.
    """
    output_path = Path(output_path)

    wb = Workbook()
    ws = wb.active
    ws.title = DPR_MASTER_DATA_SHEET

    # Header row (row 1), including blank spacer columns.
    bold = Font(bold=True)
    for col_idx, header in enumerate(DPR_MASTER_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=(header if header is not None else None))
        cell.font = bold

    # Map each real field to its 1-indexed output column.
    field_to_col = {
        header: idx
        for idx, header in enumerate(DPR_MASTER_COLUMNS, start=1)
        if header is not None
    }

    for r, row in enumerate(merged.itertuples(index=False), start=2):
        row_map = dict(zip(merged.columns, row))
        for field, col in field_to_col.items():
            value = row_map.get(field)
            if pd.isna(value):
                value = None
            if field == "date" and isinstance(value, date) and not isinstance(value, datetime):
                value = datetime.combine(value, datetime.min.time())
            ws.cell(row=r, column=col, value=value)

    # Format the date column as a date.
    if "date" in field_to_col:
        date_letter = get_column_letter(field_to_col["date"])
        for r in range(2, len(merged) + 2):
            ws[f"{date_letter}{r}"].number_format = "yyyy-mm-dd"

    # QA Flags sheet.
    qa_ws = wb.create_sheet(DPR_MASTER_QA_SHEET)
    for col_idx, header in enumerate(DPR_QA_COLUMNS, start=1):
        qa_ws.cell(row=1, column=col_idx, value=header).font = bold
    for r, flag in enumerate(qa_flags or [], start=2):
        for col_idx, header in enumerate(DPR_QA_COLUMNS, start=1):
            qa_ws.cell(row=r, column=col_idx, value=flag.get(header))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    # Atomic write: save to a temp file in the same dir, then os.replace so a
    # crash mid-save can't leave a truncated/partial master workbook.
    fd, tmp_name = tempfile.mkstemp(dir=str(output_path.parent), suffix=".xlsx.tmp")
    os.close(fd)
    try:
        wb.save(tmp_name)
        os.replace(tmp_name, str(output_path))
    except Exception:
        try:
            os.unlink(tmp_name)
        except OSError:
            pass
        raise
    logger.info(
        "Wrote DPR master %s: %d data row(s), %d QA flag(s)",
        output_path.name, len(merged), len(qa_flags or []),
    )


def write_dpr_master(
    new_records: List[Dict[str, Any]],
    output_path: Path,
    existing_master_path: Optional[Path] = None,
    qa_flags: Optional[List[Dict[str, str]]] = None,
) -> pd.DataFrame:
    """Merge *new_records* into an optional existing master and write the result.

    Convenience wrapper: merges then writes. Callers that also need the merged
    frame (e.g. for gap-flag computation or a CSV mirror) should instead call
    :func:`merge_master` once and pass the result to :func:`write_master_dataframe`.

    Returns the merged ``Data`` DataFrame (internal schema).
    """
    merged = merge_master(new_records, existing_master_path, qa_flags=qa_flags)
    write_master_dataframe(merged, output_path, qa_flags=qa_flags)
    return merged
