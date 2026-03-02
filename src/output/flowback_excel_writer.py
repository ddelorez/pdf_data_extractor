"""
Flowback Excel output writer module.
Creates a from-scratch workbook with a 3-row header matching the
OMAHA flowback data format (23 columns, sheet "Flowback Data").
"""

import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
)
from openpyxl.utils import get_column_letter

from src.config import (
    get_logger,
    FLOWBACK_COL_MAP,
    FLOWBACK_HEADER_ROW_1,
    FLOWBACK_HEADER_MERGES_ROW_1,
    FLOWBACK_HEADER_ROW_2,
    FLOWBACK_HEADER_ROW_3,
    FLOWBACK_START_ROW,
)

logger = get_logger(__name__)

# ── Style constants ────────────────────────────────────────────────────────────
_LIGHT_BLUE = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
_GRAY_FILL   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
_BOLD        = Font(bold=True)
_BOLD_ITALIC = Font(bold=True, italic=True)
_ITALIC      = Font(italic=True)

# Column-width hints  (field → chars)
_COL_WIDTHS: Dict[str, float] = {
    "Name":    25,
    "Date":    12,
    "comment": 30,
}
_DEFAULT_WIDTH = 10.0


def write_flowback_excel(
    records: List[Dict[str, Any]],
    output_path: Union[str, Path],
) -> str:
    """
    Write flowback records to a new Excel workbook.

    The workbook contains a single sheet named "Flowback Data" with a
    3-row header (group labels / field names / units) followed by one
    data row per record starting at row 4 (``FLOWBACK_START_ROW``).

    Args:
        records:     List of dicts produced by the flowback extractor.
                     Each dict may contain a ``_format`` key which is
                     silently ignored.
        output_path: Destination ``.xlsx`` file path.

    Returns:
        The resolved output path as a string.

    Raises:
        Exception: Re-raises any ``openpyxl`` save error (e.g. permission
                   denied).

    Example:
        >>> records = [{"Name": "OMAHA 12", "Date": datetime.date(2024, 1, 15),
        ...             "qo": 120}]
        >>> write_flowback_excel(records, "flowback_output.xlsx")
        'flowback_output.xlsx'
    """
    output_path = Path(output_path)
    logger.info(f"Writing flowback Excel → {output_path.name} ({len(records)} records)")

    wb = Workbook()
    ws = wb.active
    assert ws is not None, "Workbook().active must not be None"
    ws.title = "Flowback Data"

    # ── Row 1: group headers ────────────────────────────────────────────────
    for col, label in FLOWBACK_HEADER_ROW_1.items():
        cell = ws.cell(row=1, column=col, value=label if label else None)
        cell.font      = _BOLD
        cell.alignment = _CENTER
        cell.fill      = _LIGHT_BLUE

    # Merge group-header cells
    for start_col, end_col in FLOWBACK_HEADER_MERGES_ROW_1:
        ws.merge_cells(
            start_row=1, start_column=start_col,
            end_row=1,   end_column=end_col,
        )

    # ── Row 2: field name headers ───────────────────────────────────────────
    for col, label in FLOWBACK_HEADER_ROW_2.items():
        cell = ws.cell(row=2, column=col, value=label)
        cell.font      = _BOLD
        cell.alignment = _CENTER

    # ── Row 3: units ────────────────────────────────────────────────────────
    for col, unit in FLOWBACK_HEADER_ROW_3.items():
        cell = ws.cell(row=3, column=col, value=unit)
        cell.font      = _ITALIC
        cell.alignment = _CENTER
        cell.fill      = _GRAY_FILL

    # ── Rows 4+: data ──────────────────────────────────────────────────────
    # Build a reverse map: field_name → col_number (same as FLOWBACK_COL_MAP)
    field_to_col = FLOWBACK_COL_MAP  # already field → col

    for i, record in enumerate(records):
        excel_row = FLOWBACK_START_ROW + i
        # Strip internal metadata key
        rec = {k: v for k, v in record.items() if k != "_format"}

        for field_name, col_number in field_to_col.items():
            value = rec.get(field_name)

            if value is None:
                continue

            # Format date objects as "M/D/YYYY"
            if field_name == "Date" and isinstance(value, (datetime.date, datetime.datetime)):
                value = f"{value.month}/{value.day}/{value.year}"

            ws.cell(row=excel_row, column=col_number, value=value)

    # ── Column widths ───────────────────────────────────────────────────────
    for field_name, col_number in field_to_col.items():
        col_letter = get_column_letter(col_number)
        width = _COL_WIDTHS.get(field_name, _DEFAULT_WIDTH)
        ws.column_dimensions[col_letter].width = width

    # ── Freeze panes (keep header rows visible) ─────────────────────────────
    # freeze_panes accepts a cell address string ("A4" keeps rows 1-3 visible)
    ws.freeze_panes = f"A{FLOWBACK_START_ROW}"

    # ── Save ────────────────────────────────────────────────────────────────
    try:
        wb.save(output_path)
        logger.info(f"✅ Flowback Excel written: {output_path}")
    except Exception as exc:
        logger.error(f"Failed to save flowback Excel: {exc}")
        raise

    return str(output_path)
