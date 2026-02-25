"""
Excel output writer module.
Handles template loading and populating with extracted data.
"""

from datetime import datetime
from pathlib import Path
from typing import Union, Optional, Dict, Any

import pandas as pd
from openpyxl import load_workbook

from src.config import (
    get_logger,
    START_ROW,
    COL_MAP,
    OUTPUT_XLSX,
)

logger = get_logger(__name__)


def write_excel(
    df: pd.DataFrame,
    template_path: Union[str, Path],
    output_path: Optional[Union[str, Path]] = None,
) -> Path:
    """
    Write extracted data to Excel file using template.
    
    Process:
    1. Load existing Excel template
    2. Clear old data rows (from START_ROW onward)
    3. Populate cells from DataFrame using COL_MAP
    4. Preserve formatting and formulas in template
    5. Save to output file
    
    Args:
        df: Pandas DataFrame with extracted records
        template_path: Path to Excel template file
        output_path: Where to save output (default: config.OUTPUT_XLSX)
    
    Returns:
        Path to written Excel file
    
    Raises:
        FileNotFoundError: If template file not found
        Exception: If Excel writing fails (e.g., permission denied)
    
    Example:
        >>> df = pd.DataFrame([
        ...     {"Well": "W-01", "Date": date(2024, 1, 15), "qo": 100}
        ... ])
        >>> output_path = write_excel(df, "template.xlsx")
        >>> print(f"Written to {output_path}")
    """
    template_path = Path(template_path)
    output_path = Path(output_path or OUTPUT_XLSX)
    
    # Validate template exists
    if not template_path.exists():
        logger.error(f"Template file not found: {template_path}")
        raise FileNotFoundError(f"Template file not found: {template_path}")
    
    # Load workbook
    logger.info(f"Loading template: {template_path.name}")
    try:
        wb = load_workbook(template_path)
        ws = wb.active
    except Exception as e:
        logger.error(f"Failed to load template: {e}")
        raise
    
    # Clear existing data rows
    logger.debug(f"Clearing data rows from {START_ROW} to {ws.max_row}")
    for row in ws.iter_rows(min_row=START_ROW, max_row=ws.max_row):
        for cell in row:
            cell.value = None
    
    # Write data from DataFrame
    logger.info(f"Writing {len(df)} records to Excel")
    
    for i, row_data in enumerate(df.itertuples(index=False)):
        excel_row = START_ROW + i
        
        for field_name, col_number in COL_MAP.items():
            # Get value from dataframe
            if hasattr(row_data, field_name):
                value = getattr(row_data, field_name)
                
                # Convert date to datetime for Excel compatibility
                if field_name == "Date" and value is not None:
                    if not isinstance(value, datetime):
                        # Convert date object to datetime
                        try:
                            value = datetime.combine(value, datetime.min.time())
                        except Exception as e:
                            logger.warning(f"Failed to convert date to datetime: {e}")
                            value = None
                
                ws.cell(row=excel_row, column=col_number, value=value)
    
    # Save workbook
    logger.info(f"Saving Excel file: {output_path.name}")
    try:
        wb.save(output_path)
        logger.info(f"✅ Excel file written: {output_path}")
    except Exception as e:
        logger.error(f"Failed to save Excel file: {e}")
        raise
    
    return output_path


def get_excel_summary(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Generate summary statistics from extracted data.
    
    Args:
        df: Pandas DataFrame with extracted records
    
    Returns:
        Dictionary with summary statistics
    
    Example:
        >>> summary = get_excel_summary(df)
        >>> print(f"Total oil: {summary['total_oil']} barrels")
    """
    summary = {
        "total_records": len(df),
        "unique_wells": df['Well'].nunique() if not df.empty else 0,
        "date_range": None,
        "total_oil": 0,
        "total_gas": 0,
        "total_water": 0,
    }
    
    if not df.empty:
        # Date range
        if "Date" in df.columns:
            valid_dates = df["Date"].dropna()
            if not valid_dates.empty:
                summary["date_range"] = {
                    "start": valid_dates.min(),
                    "end": valid_dates.max(),
                }
        
        # Production totals
        if "qo" in df.columns:
            summary["total_oil"] = int(df["qo"].sum()) if df["qo"].notna().any() else 0
        
        if "qg" in df.columns:
            summary["total_gas"] = int(df["qg"].sum()) if df["qg"].notna().any() else 0
        
        if "qw" in df.columns:
            summary["total_water"] = int(df["qw"].sum()) if df["qw"].notna().any() else 0
    
    return summary
