"""Synthetic builders for the DPR Excel -> Excel test suite.

These write *small* Walter-Oil-shaped ``.xlsx`` workbooks with openpyxl so the
tests never depend on the gitignored ``Reference-files/`` directory (which CI
does not have). The geometry mirrors the real partner layout closely enough to
exercise :mod:`src.core.excel_format_detector`,
:mod:`src.core.excel_dpr_extraction`, :mod:`src.data.dpr_qa`, and
:mod:`src.output.dpr_master_writer`.

Real geometry reproduced (per daily sheet):
  * E3 = "EW 834-A Platform"
  * L3 = "Report Date",     N3 = report datetime
  * L4 = "Production Date", N4 = production datetime  (the authoritative date)
  * row 8 = header row (Block / Well Number / FTP / BHP / ... / Notes)
  * well rows 11-19 (first group) and 21-23 (second group); row 20 is a blank gap
  * D24 = "PLATFORM SALES….." (footer, must not be read as a well row)

Two non-daily sheets (" Summary" and "Meter totals") are always present and must
be ignored by the extractor/detector.
"""

from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from openpyxl import Workbook

# Default well ids, in the order they occupy the two well-row ranges below.
# 9 wells in rows 11-19, then 3 in rows 21-23 -> 12 wells total.
DEFAULT_WELLS = [f"A-{i}" for i in range(1, 10)] + ["SS001", "SS002", "SS003"]

# Inclusive 1-indexed well-row ranges (row 20 intentionally skipped).
WELL_ROW_RANGES = [(11, 19), (21, 23)]

# output field -> source column letter (mirrors config DPR_EXCEL_FORMATS columns)
FIELD_TO_COL = {
    "FTP": "D",
    "BHP": "E",
    "Choke Size": "H",
    "Daily Gas": "I",
    "Daily Oil": "J",
    "Daily Water": "K",
}

_HEADER_ROW_8 = {
    "B8": "Block",
    "C8": "Well Number",
    "D8": "FTP",
    "E8": "BHP",
    "F8": "BHT",
    "G8": "FLP FA-3",
    "H8": "Choke #2",
    "I8": "Est Allocated Daily Gas Vol   MCFPD",
    "J8": "Est Allocated Daily Oil Vol   BOPD",
    "K8": "Est Allocated Daily Wtr Vol    BWPD",
    "L8": "SITP",
    "M8": "Daily DT Hrs",
    "N8": "Notes & Reason for DT",
}


def _well_rows():
    """Yield the 1-indexed worksheet rows that hold well data, in order."""
    for start, end in WELL_ROW_RANGES:
        for row in range(start, end + 1):
            yield row


def _default_values(well_index: int) -> Dict[str, Any]:
    """Deterministic per-well numeric defaults (overridable via *values*)."""
    return {
        "FTP": 500.0 + well_index,
        "BHP": 2000.0 + well_index,
        "Choke Size": 20.0 + well_index,
        "Daily Gas": 1000.0 + well_index,
        "Daily Oil": 100.0 + well_index,
        "Daily Water": 10.0 + well_index,
    }


def build_dpr_workbook(
    path: Union[str, Path],
    month_dates_by_sheet: Dict[str, Union[date, datetime]],
    wells: Optional[List[str]] = None,
    values: Optional[Dict[str, Dict[str, Dict[str, Any]]]] = None,
    report_dates_by_sheet: Optional[Dict[str, Union[date, datetime]]] = None,
    blank_wells: Optional[List[str]] = None,
) -> Path:
    """Write a synthetic Walter Oil DPR workbook to *path*.

    Args:
        path: Destination ``.xlsx``.
        month_dates_by_sheet: Maps daily-sheet name (an integer string, e.g.
            ``"1"``) to the datetime written to that sheet's N4 (Production
            Date). Pass an out-of-month date (e.g. ``date(2023, 12, 31)``) on one
            sheet to exercise the QA date check.
        wells: Well ids to lay down the two row ranges (default: 12 wells).
        values: Optional overrides ``{sheet: {well: {field: raw_value}}}``. The
            raw value is written verbatim, so tests can inject sentinels
            (``"S/I"``, ``"NA"``), a genuine ``0``, or a full-precision float.
        report_dates_by_sheet: Optional N3 (Report Date) per sheet; defaults to
            the same value as N4.
        blank_wells: Well ids to leave with an empty col-C cell (skipped rows).

    Returns:
        The path written.
    """
    path = Path(path)
    wells = wells if wells is not None else list(DEFAULT_WELLS)
    values = values or {}
    report_dates_by_sheet = report_dates_by_sheet or {}
    blank_wells = set(blank_wells or [])

    wb = Workbook()

    # --- non-daily sheets that must be ignored ---
    summary = wb.active
    summary.title = " Summary"
    summary["A1"] = "Monthly Summary"
    meters = wb.create_sheet("Meter totals")
    meters["A1"] = "Meter totals"

    # --- daily sheets ---
    for sheet_name, prod_date in month_dates_by_sheet.items():
        ws = wb.create_sheet(sheet_name)
        ws["E3"] = "EW 834-A Platform"
        ws["L3"] = "Report Date"
        ws["N3"] = report_dates_by_sheet.get(sheet_name, prod_date)
        ws["L4"] = "Production Date"
        ws["N4"] = prod_date

        for coord, text in _HEADER_ROW_8.items():
            ws[coord] = text

        sheet_overrides = values.get(sheet_name, {})
        rows = list(_well_rows())
        for well_index, (well_id, row) in enumerate(zip(wells, rows)):
            if well_id in blank_wells:
                continue  # leave col-C blank -> extractor skips the row
            ws[f"C{row}"] = well_id
            cell_values = _default_values(well_index)
            cell_values.update(sheet_overrides.get(well_id, {}))
            for field, col in FIELD_TO_COL.items():
                ws[f"{col}{row}"] = cell_values.get(field)

        ws["D24"] = "PLATFORM SALES….."

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path


# Master workbook column order (mirrors config DPR_MASTER_COLUMNS; None = spacer).
_MASTER_COLUMNS = [
    "well",
    "date",
    "Daily Oil",
    "Daily Gas",
    "Daily Water",
    "BHP",
    "FTP",
    None,
    None,
    None,
    "Choke Size",
]


def build_master_workbook(
    path: Union[str, Path],
    rows: List[Dict[str, Any]],
) -> Path:
    """Write a minimal existing "master" workbook (``Data`` sheet only).

    Header row matches :data:`_MASTER_COLUMNS` with the three blank spacer
    columns written as empty cells; ``date`` lands in column 2 and
    ``Choke Size`` in column 11. Each row dict supplies ``well``, ``date`` and
    any of the real data fields.

    Returns the path written.
    """
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    for col_idx, header in enumerate(_MASTER_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=(header if header is not None else None))

    field_to_col = {
        header: idx
        for idx, header in enumerate(_MASTER_COLUMNS, start=1)
        if header is not None
    }

    for r, row in enumerate(rows, start=2):
        for field, col in field_to_col.items():
            value = row.get(field)
            if field == "date" and isinstance(value, date) and not isinstance(value, datetime):
                value = datetime.combine(value, datetime.min.time())
            cell = ws.cell(row=r, column=col, value=value)
            if field == "date" and value is not None:
                cell.number_format = "yyyy-mm-dd"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path


def build_plain_workbook(path: Union[str, Path]) -> Path:
    """Write an unrelated one-sheet workbook (should classify as UNKNOWN)."""
    path = Path(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "hello"
    ws["B1"] = "world"
    ws["A2"] = 1
    ws["B2"] = 2
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return path
