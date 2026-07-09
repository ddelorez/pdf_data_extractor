"""Integration tests for the DPR Excel -> Excel pipeline at the service layer.

Exercises ExtractionService end-to-end for .xlsx input: modality routing in
process_job, _process_dpr_job (classification, extract, append+dedup, master +
CSV output), and the .xlsx upload validation added to submit_files. Uses the
synthetic builders in tests/backend/fixtures/dpr_builder.py so nothing depends on
the gitignored Reference-files/ directory.
"""

import io
from datetime import date, datetime

import pytest
from openpyxl import load_workbook, Workbook
from werkzeug.datastructures import FileStorage

import services.extraction_service as svc
from services.extraction_service import (
    ExtractionService,
    ProcessingJob,
    ProcessingError,
    NonPdfFileError,
    _looks_like_xlsx,
    _looks_like_pdf,
)
from tests.backend.fixtures.dpr_builder import (
    build_dpr_workbook,
    build_master_workbook,
    build_plain_workbook,
)


@pytest.fixture
def service(tmp_path):
    up = tmp_path / "uploads"
    out = tmp_path / "outputs"
    tpl = tmp_path / "templates"
    return ExtractionService(str(up), str(tpl), str(out))


def _register_job_with_files(service, file_paths):
    """Create a PENDING job whose folder holds the given .xlsx files."""
    import uuid
    job_id = str(uuid.uuid4())
    job = ProcessingJob(job_id, str(service.upload_folder))
    for i, src in enumerate(file_paths):
        dest = job.job_folder / f"file_{i}.xlsx"
        dest.write_bytes(src.read_bytes())
        job.add_file(dest)
    service.jobs[job_id] = job
    service._persist_job(job)
    return job_id


@pytest.mark.integration
class TestDprServicePipeline:
    def test_single_workbook_conversion(self, service, tmp_path):
        raw = build_dpr_workbook(
            tmp_path / "april.xlsx",
            month_dates_by_sheet={"2": datetime(2026, 4, 1), "3": datetime(2026, 4, 2)},
        )
        job_id = _register_job_with_files(service, [raw])

        result = service.process_job(job_id)

        assert result["status"] == "success"
        # 12 wells x 2 days
        assert result["records"] == 24
        assert result["unique_wells"] == 12

        xlsx_out = service.output_folder / f"{job_id}_output.xlsx"
        csv_out = service.output_folder / f"{job_id}_output.csv"
        assert xlsx_out.exists() and csv_out.exists()

        wb = load_workbook(xlsx_out)
        assert wb.sheetnames == ["Data", "QA Flags"]
        data = wb["Data"]
        # header row includes blank spacer columns; Choke Size in col 11
        assert data.cell(row=1, column=1).value == "well"
        assert data.cell(row=1, column=11).value == "Choke Size"

    def test_append_to_existing_master_newest_wins(self, service, tmp_path):
        # Existing master already has A-1 @ 2026-04-01 with a stale oil value.
        master = build_master_workbook(
            tmp_path / "master.xlsx",
            rows=[
                {"well": "A-1", "date": date(2026, 4, 1), "Daily Oil": 1.0,
                 "Daily Gas": 1.0, "Daily Water": 1.0, "BHP": 1.0, "FTP": 1.0,
                 "Choke Size": 1.0},
            ],
        )
        # New raw workbook re-reports 2026-04-01 (overlap) plus 2026-04-02.
        raw = build_dpr_workbook(
            tmp_path / "april.xlsx",
            month_dates_by_sheet={"2": datetime(2026, 4, 1), "3": datetime(2026, 4, 2)},
            values={"2": {"A-1": {"Daily Oil": 999.0}}},
        )
        job_id = _register_job_with_files(service, [raw, master])

        result = service.process_job(job_id)
        assert result["status"] == "success"

        # 12 wells x 2 days = 24; the stale master row is overwritten, not added.
        assert result["records"] == 24

        wb = load_workbook(service.output_folder / f"{job_id}_output.xlsx")
        data = wb["Data"]
        # find A-1 @ 2026-04-01 and confirm newest (999.0) won
        found = None
        for r in range(2, data.max_row + 1):
            well = data.cell(row=r, column=1).value
            dval = data.cell(row=r, column=2).value
            d = dval.date() if hasattr(dval, "date") else dval
            if well == "A-1" and d == date(2026, 4, 1):
                found = data.cell(row=r, column=3).value  # Daily Oil
        assert found == 999.0

    def test_qa_flag_written_for_bad_date(self, service, tmp_path):
        raw = build_dpr_workbook(
            tmp_path / "april.xlsx",
            month_dates_by_sheet={
                "2": datetime(2026, 4, 1),
                "3": datetime(2026, 4, 2),
                "1": datetime(2023, 12, 31),  # stale template default -> flagged
            },
        )
        job_id = _register_job_with_files(service, [raw])
        service.process_job(job_id)

        wb = load_workbook(service.output_folder / f"{job_id}_output.xlsx")
        qa = wb["QA Flags"]
        concerns = [qa.cell(row=r, column=3).value for r in range(2, qa.max_row + 1)]
        assert any(c and "2023-12-31" in c for c in concerns)

    def test_no_master_supplied_adds_qa_flag(self, service, tmp_path):
        # HIGH-3: a raw-only batch succeeds but must warn that the output is
        # history-free so the user doesn't adopt it as the full master.
        raw = build_dpr_workbook(
            tmp_path / "april.xlsx",
            month_dates_by_sheet={"2": datetime(2026, 4, 1)},
        )
        job_id = _register_job_with_files(service, [raw])
        service.process_job(job_id)

        wb = load_workbook(service.output_folder / f"{job_id}_output.xlsx")
        qa = wb["QA Flags"]
        concerns = [qa.cell(row=r, column=3).value for r in range(2, qa.max_row + 1)]
        assert any(c and "No existing master" in c for c in concerns)

    def test_decompression_bomb_guard(self, service, tmp_path, monkeypatch):
        # MED-3: an .xlsx whose uncompressed size exceeds the cap is refused
        # before any parse. Force a tiny cap so a normal file trips it.
        monkeypatch.setattr(svc, "MAX_XLSX_UNCOMPRESSED_BYTES", 10)
        raw = build_dpr_workbook(
            tmp_path / "april.xlsx",
            month_dates_by_sheet={"2": datetime(2026, 4, 1)},
        )
        job_id = _register_job_with_files(service, [raw])
        with pytest.raises(ProcessingError, match="exceeds"):
            service.process_job(job_id)

    def test_unrecognized_workbook_errors(self, service, tmp_path):
        plain = build_plain_workbook(tmp_path / "plain.xlsx")
        job_id = _register_job_with_files(service, [plain])
        with pytest.raises(ProcessingError, match="Unrecognized"):
            service.process_job(job_id)

    def test_master_only_no_raw_errors(self, service, tmp_path):
        master = build_master_workbook(
            tmp_path / "master.xlsx",
            rows=[{"well": "A-1", "date": date(2026, 4, 1), "Daily Oil": 1.0,
                   "Daily Gas": 1.0, "Daily Water": 1.0, "BHP": 1.0, "FTP": 1.0,
                   "Choke Size": 1.0}],
        )
        job_id = _register_job_with_files(service, [master])
        with pytest.raises(ProcessingError, match="No DPR workbook"):
            service.process_job(job_id)


@pytest.mark.integration
class TestUploadValidation:
    def test_magic_byte_helpers(self):
        # A real .xlsx (zip) begins with PK\x03\x04.
        buf = io.BytesIO()
        Workbook().save(buf)
        header = buf.getvalue()[:1024]
        assert _looks_like_xlsx(header) is True
        assert _looks_like_pdf(header) is False
        assert _looks_like_pdf(b"%PDF-1.7 ...") is True
        assert _looks_like_xlsx(b"%PDF-1.7 ...") is False

    def test_submit_accepts_xlsx(self, service, tmp_path, monkeypatch):
        # Prevent the background pool from auto-processing during this
        # validation-only test.
        monkeypatch.setattr(svc._processing_executor, "submit", lambda *a, **k: None)

        buf = io.BytesIO()
        Workbook().save(buf)
        buf.seek(0)
        fs = FileStorage(stream=buf, filename="book.xlsx",
                         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        job_id = service.submit_files([fs])

        job = service.jobs[job_id]
        assert len(job.files_submitted) == 1
        assert job.files_submitted[0].suffix == ".xlsx"

    def test_submit_rejects_unsupported_extension(self, service, monkeypatch):
        monkeypatch.setattr(svc._processing_executor, "submit", lambda *a, **k: None)
        fs = FileStorage(stream=io.BytesIO(b"hello"), filename="notes.txt")
        with pytest.raises(NonPdfFileError):
            service.submit_files([fs])

    def test_submit_rejects_xlsx_with_bad_magic(self, service, monkeypatch):
        monkeypatch.setattr(svc._processing_executor, "submit", lambda *a, **k: None)
        fs = FileStorage(stream=io.BytesIO(b"not really a zip"), filename="fake.xlsx")
        with pytest.raises(NonPdfFileError):
            service.submit_files([fs])
