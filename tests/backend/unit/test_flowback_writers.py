"""
Unit tests for the flowback output writers (tabular_flowback format).
These lock the current output structure (3-row header + data rows) so a future
dependency bump can be diffed against known behavior.
"""

import csv
from datetime import date

import pytest
from openpyxl import load_workbook

from src.config import (
    FLOWBACK_COL_MAP,
    FLOWBACK_HEADER_ROW_2,
    FLOWBACK_HEADER_ROW_3,
    FLOWBACK_START_ROW,
)
from src.output.flowback_csv_writer import write_flowback_csv
from src.output.flowback_excel_writer import write_flowback_excel


@pytest.fixture
def flowback_records():
    """Two flowback records; the second exercises None-skipping."""
    return [
        {
            "_format": "tabular_flowback",
            "Name": "OMAHA 12",
            "Well": "OMAHA 12",
            "Date": date(2024, 1, 15),
            "qo": 120,
            "qg": 5400,
            "qw": 30,
            "ptubing": 2100,
            "pcasing": 3300,
            "days_on": 28,
            "comment": "stable",
        },
        {
            "_format": "tabular_flowback",
            "Name": "OMAHA 13",
            "Well": "OMAHA 13",
            "Date": date(2024, 1, 16),
            "qo": 95,
            "qg": None,          # None must be written as empty, not "None"
            "qw": 18,
            "comment": None,
        },
    ]


class TestWriteFlowbackCsv:
    def test_returns_path_and_creates_file(self, flowback_records, tmp_path):
        out = tmp_path / "fb.csv"
        result = write_flowback_csv(flowback_records, out)
        assert result == str(out)
        assert out.exists()

    def test_three_header_rows_then_data(self, flowback_records, tmp_path):
        out = tmp_path / "fb.csv"
        write_flowback_csv(flowback_records, out)
        with open(out, newline="", encoding="utf-8") as fh:
            rows = list(csv.reader(fh))

        # 3 header rows + 2 data rows
        assert len(rows) == 5
        num_cols = max(FLOWBACK_COL_MAP.values())
        assert all(len(r) == num_cols for r in rows)

        # Row 2 = field names, Row 3 = units (per config maps)
        assert rows[1][0] == FLOWBACK_HEADER_ROW_2[1]   # 'Name'
        assert rows[2][2] == FLOWBACK_HEADER_ROW_3[3]   # qo units 'STB/d'

    def test_data_values_and_date_format(self, flowback_records, tmp_path):
        out = tmp_path / "fb.csv"
        write_flowback_csv(flowback_records, out)
        with open(out, newline="", encoding="utf-8") as fh:
            rows = list(csv.reader(fh))

        first = rows[FLOWBACK_START_ROW - 1]  # first data row (row index 3)
        assert first[0] == "OMAHA 12"                    # Name (col 1)
        assert first[1] == "1/15/2024"                   # Date M/D/YYYY (col 2)
        assert first[2] == "120"                         # qo (col 3)

    def test_none_values_written_empty(self, flowback_records, tmp_path):
        out = tmp_path / "fb.csv"
        write_flowback_csv(flowback_records, out)
        with open(out, newline="", encoding="utf-8") as fh:
            rows = list(csv.reader(fh))
        second = rows[FLOWBACK_START_ROW]  # second data row
        # qg is column 4 -> index 3; None should be empty string, never "None"
        assert second[3] == ""

    def test_format_tag_not_emitted_as_column(self, flowback_records, tmp_path):
        out = tmp_path / "fb.csv"
        write_flowback_csv(flowback_records, out)
        content = out.read_text(encoding="utf-8")
        assert "_format" not in content
        assert "tabular_flowback" not in content


class TestWriteFlowbackExcel:
    def test_returns_path_and_creates_file(self, flowback_records, tmp_path):
        out = tmp_path / "fb.xlsx"
        result = write_flowback_excel(flowback_records, out)
        assert result == str(out)
        assert out.exists()

    def test_sheet_and_header_and_data(self, flowback_records, tmp_path):
        out = tmp_path / "fb.xlsx"
        write_flowback_excel(flowback_records, out)
        wb = load_workbook(out)
        ws = wb["Flowback Data"]

        # Field-name header row (row 2) and unit row (row 3)
        assert ws.cell(row=2, column=1).value == FLOWBACK_HEADER_ROW_2[1]  # 'Name'
        assert ws.cell(row=3, column=3).value == FLOWBACK_HEADER_ROW_3[3]  # 'STB/d'

        # First data row at FLOWBACK_START_ROW
        assert ws.cell(row=FLOWBACK_START_ROW, column=1).value == "OMAHA 12"
        assert ws.cell(row=FLOWBACK_START_ROW, column=2).value == "1/15/2024"
        assert ws.cell(row=FLOWBACK_START_ROW, column=3).value == 120

    def test_empty_records_writes_headers_only(self, tmp_path):
        out = tmp_path / "fb_empty.xlsx"
        write_flowback_excel([], out)
        wb = load_workbook(out)
        ws = wb["Flowback Data"]
        # Header present, no data row
        assert ws.cell(row=2, column=1).value == FLOWBACK_HEADER_ROW_2[1]
        assert ws.cell(row=FLOWBACK_START_ROW, column=1).value is None
