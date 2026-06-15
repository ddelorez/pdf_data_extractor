"""
Additional unit tests for the SOR output writers, covering gaps:
- excel_writer.get_excel_summary
- excel_writer.write_excel template (3-arg) + missing-template fallback + 2-arg
- csv_writer.write_csv_with_formatting (date column formatting, no index)
"""

from datetime import date, datetime

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from src.config import START_ROW, COL_MAP
from src.output.excel_writer import write_excel, get_excel_summary
from src.output.csv_writer import write_csv_with_formatting


@pytest.fixture
def df(sample_records):
    return pd.DataFrame(sample_records)


class TestGetExcelSummary:
    def test_empty_dataframe(self):
        summary = get_excel_summary(pd.DataFrame())
        assert summary["total_records"] == 0
        assert summary["unique_wells"] == 0
        assert summary["date_range"] is None
        assert summary["total_oil"] == 0
        assert summary["total_gas"] == 0
        assert summary["total_water"] == 0

    def test_populated_dataframe(self, df, sample_records):
        summary = get_excel_summary(df)
        assert summary["total_records"] == len(sample_records)
        assert summary["unique_wells"] == 2  # HORIZON + WILDCAT
        assert summary["total_oil"] == sum(r["qo"] for r in sample_records)
        assert summary["total_gas"] == sum(r["qg"] for r in sample_records)
        assert summary["date_range"]["start"] == date(2024, 1, 15)
        assert summary["date_range"]["end"] == date(2024, 1, 17)


class TestWriteExcelTemplate:
    def _make_template(self, path):
        """A minimal template with a header row and stale data to be cleared."""
        wb = Workbook()
        ws = wb.active
        ws.cell(row=START_ROW - 1, column=1, value="Well")  # header above data
        ws.cell(row=START_ROW, column=1, value="STALE")     # old data to clear
        ws.cell(row=START_ROW + 1, column=1, value="STALE2")
        wb.save(path)
        return path

    def test_three_arg_template_form_places_data_by_col_map(self, df, tmp_path):
        template = self._make_template(tmp_path / "template.xlsx")
        out = tmp_path / "out.xlsx"
        result = write_excel(df, template, out)
        assert result == out

        ws = load_workbook(out).active
        # First data record lands at START_ROW, in COL_MAP positions.
        first = df.iloc[0]
        assert ws.cell(row=START_ROW, column=COL_MAP["Well"]).value == first["Well"]
        assert ws.cell(row=START_ROW, column=COL_MAP["qo"]).value == first["qo"]
        # Date converted to a datetime for Excel
        assert isinstance(ws.cell(row=START_ROW, column=COL_MAP["Date"]).value, datetime)
        # Stale row beyond the new data was cleared
        last_row = START_ROW + len(df)
        assert ws.cell(row=last_row, column=1).value is None

    def test_missing_template_falls_back_to_plain_sheet(self, df, tmp_path):
        out = tmp_path / "out.xlsx"
        write_excel(df, tmp_path / "does_not_exist.xlsx", out)
        # Plain sheet is round-trippable via pandas (headers in row 1)
        roundtrip = pd.read_excel(out)
        assert list(roundtrip.columns)[:2] == ["Well", "Date"]
        assert len(roundtrip) == len(df)

    def test_two_arg_form_writes_plain_sheet(self, df, tmp_path):
        out = tmp_path / "plain.xlsx"
        result = write_excel(df, out)  # second positional arg is the destination
        assert result == out
        roundtrip = pd.read_excel(out)
        assert len(roundtrip) == len(df)


class TestWriteCsvWithFormatting:
    def test_datetime_column_formatted_iso_no_index(self, tmp_path):
        frame = pd.DataFrame({
            "Well": ["W-01", "W-01"],
            "Date": pd.to_datetime(["2024-01-15", "2024-01-16"]),
            "qo": [100, 110],
        })
        out = tmp_path / "fmt.csv"
        result = write_csv_with_formatting(frame, out)
        assert result == out

        lines = out.read_text(encoding="utf-8").splitlines()
        assert lines[0] == "Well,Date,qo"          # header, no leading index column
        assert lines[1] == "W-01,2024-01-15,100"   # date as YYYY-MM-DD, not a timestamp

    def test_requires_dataframe_not_list(self, sample_records, tmp_path):
        # Unlike write_csv, write_csv_with_formatting expects a DataFrame.
        with pytest.raises(AttributeError):
            write_csv_with_formatting(sample_records, tmp_path / "fmt2.csv")
