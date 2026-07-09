"""Unit tests for src/output/dpr_master_writer.py (merge + write + load)."""

from datetime import date, datetime

import pytest
from openpyxl import load_workbook

from openpyxl import load_workbook as _load_wb

from src.config import DPR_MASTER_COLUMNS, DPR_MASTER_DATA_SHEET
from src.output.dpr_master_writer import (
    merge_master,
    write_dpr_master,
    load_existing_master,
    MasterReadError,
)
from tests.backend.fixtures.dpr_builder import (
    build_master_workbook,
    build_plain_workbook,
)


def _rec(well, d, **fields):
    base = {"well": well, "date": d}
    base.update(fields)
    return base


class TestMergeMaster:
    def test_dedup_keeps_newest_incoming(self, tmp_path):
        existing = build_master_workbook(
            tmp_path / "master.xlsx",
            [_rec("A-1", date(2026, 4, 1), **{"Daily Oil": 1.0})],
        )
        new_records = [_rec("A-1", date(2026, 4, 1), **{"Daily Oil": 2.0})]

        df = merge_master(new_records, existing_master_path=existing)
        assert len(df) == 1
        assert df.iloc[0]["well"] == "A-1"
        assert df.iloc[0]["Daily Oil"] == 2.0

    def test_sort_order_date_then_well(self):
        new_records = [
            _rec("B-1", date(2026, 4, 2)),
            _rec("A-1", date(2026, 4, 2)),
            _rec("Z-9", date(2026, 4, 1)),
        ]
        df = merge_master(new_records)
        order = list(zip(df["date"], df["well"]))
        assert order == [
            (date(2026, 4, 1), "Z-9"),
            (date(2026, 4, 2), "A-1"),
            (date(2026, 4, 2), "B-1"),
        ]

    def test_no_existing_returns_new_only(self):
        df = merge_master([_rec("A-1", date(2026, 4, 1), **{"Daily Oil": 5.0})])
        assert len(df) == 1
        assert df.iloc[0]["Daily Oil"] == 5.0


class TestLoadExistingMasterSafety:
    def test_unreadable_master_raises(self, tmp_path):
        # HIGH-2: a workbook with no Data sheet must fail loud, not silently
        # return an empty master (which would erase history).
        plain = build_plain_workbook(tmp_path / "plain.xlsx")
        with pytest.raises(MasterReadError):
            load_existing_master(plain)

    def test_bad_date_row_flagged_not_silent(self, tmp_path):
        # HIGH-2: rows whose date can't be parsed are excluded, but the loss is
        # surfaced as a QA flag rather than dropped silently.
        master = build_master_workbook(
            tmp_path / "master.xlsx",
            [
                _rec("A-1", date(2026, 4, 1), **{"Daily Oil": 1.0}),
                _rec("A-2", date(2026, 4, 1), **{"Daily Oil": 2.0}),
            ],
        )
        # Corrupt the second data row's date cell (col 2, row 3) to text.
        wb = _load_wb(master)
        wb[DPR_MASTER_DATA_SHEET].cell(row=3, column=2, value="not-a-date")
        wb.save(str(master))

        flags = []
        df = load_existing_master(master, qa_flags=flags)
        assert len(df) == 1  # the bad row is excluded
        assert any("unparseable" in f["Concern"] for f in flags)


class TestWriteDprMaster:
    def test_sheets_and_header(self, tmp_path):
        out = tmp_path / "out.xlsx"
        qa_flags = [
            {"Workbook": "April.xlsx", "Sheet": "3", "Concern": "N4 date was 2023-12-31"}
        ]
        write_dpr_master(
            [_rec("A-1", date(2026, 4, 1), **{"Daily Oil": 1.0, "Choke Size": 20.0})],
            out,
            qa_flags=qa_flags,
        )

        wb = load_workbook(out)
        assert wb.sheetnames == ["Data", "QA Flags"]

        data = wb["Data"]
        header = [data.cell(row=1, column=c).value for c in range(1, len(DPR_MASTER_COLUMNS) + 1)]
        assert header == DPR_MASTER_COLUMNS  # spacers are None (blank cells)
        # Choke Size occupies column 11.
        assert data.cell(row=1, column=11).value == "Choke Size"
        assert data.cell(row=2, column=11).value == 20.0

        qa = wb["QA Flags"]
        assert [qa.cell(row=1, column=c).value for c in range(1, 4)] == [
            "Workbook", "Sheet", "Concern",
        ]
        assert qa.cell(row=2, column=1).value == "April.xlsx"
        assert qa.cell(row=2, column=3).value == "N4 date was 2023-12-31"

    def test_date_in_column_2(self, tmp_path):
        out = tmp_path / "out.xlsx"
        write_dpr_master([_rec("A-1", date(2026, 4, 5))], out)
        wb = load_workbook(out)
        cell = wb["Data"].cell(row=2, column=2).value
        assert isinstance(cell, datetime)
        assert cell.date() == date(2026, 4, 5)

    def test_round_trip_load_existing_master(self, tmp_path):
        out = tmp_path / "out.xlsx"
        records = [
            _rec("A-1", date(2026, 4, 1), **{"Daily Oil": 100.0, "Choke Size": 20.0}),
            _rec("SS001", date(2026, 4, 2), **{"Daily Oil": 2041.95062750134}),
        ]
        write_dpr_master(records, out)

        df = load_existing_master(out)
        assert len(df) == 2
        by_well = {r["well"]: r for _, r in df.iterrows()}
        assert by_well["A-1"]["date"] == date(2026, 4, 1)
        assert by_well["A-1"]["Daily Oil"] == 100.0
        assert by_well["SS001"]["Daily Oil"] == 2041.95062750134
